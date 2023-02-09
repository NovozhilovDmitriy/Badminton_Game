import sqlite3
from sqlite3 import Error
from main import read_config
import openpyxl as xl


def create_table_game():
    conn = create_connection()
    games_table = '''CREATE TABLE IF NOT EXISTS games (
                    id INTEGER PRIMARY KEY,
                    date TEXT NOT NULL,
                    player1 TEXT NOT NULL,
                    player2 TEXT NOT NULL,
                    player3 TEXT NOT NULL,
                    player4 TEXT NOT NULL,
                    score1 INTEGER NOT NULL,
                    score2 INTEGER NOT NULL);'''
    players_score = '''CREATE TABLE IF NOT EXISTS players (
                    id INTEGER PRIMARY KEY,
                    name TEXT NOT NULL,
                    score REAL NOT NULL,
                    daily_score REAL NOT NULL);'''
    cursor = conn.cursor()
    cursor.execute(games_table)
    cursor.execute(players_score)
    conn.commit()
    cursor.fetchall()
    #print('Table games and table score created (if not exist)')
    cursor.close()

def create_stat_table():
    conn = create_connection()
    stat_table = '''CREATE TABLE IF NOT EXISTS stat (
                        id INTEGER PRIMARY KEY,
                        date TEXT NOT NULL,
                        player1 TEXT NOT NULL,
                        pl1_sum_daily_score REAL NOT NULL,
                        pl1_current_score REAL NOT NULL,
                        pl1_play_games INTEGER NOT NULL,
                        player2 TEXT NOT NULL,
                        pl2_sum_daily_score REAL NOT NULL,
                        pl2_current_score REAL NOT NULL,
                        pl2_play_games INTEGER NOT NULL,
                        score1 INTEGER NOT NULL,
                        win_lose1 INTEGER NOT NULL,
                        max_min1 INTEGER NOT NULL,
                        more10_1 INTEGER NOT NULL,
                        pair1_avr_score REAL NOT NULL,
                        pair1_Ea REAL NOT NULL,
                        pair1_Sa REAL NOT NULL,
                        pair1_Ra REAL NOT NULL,
                        player3 TEXT NOT NULL,
                        pl3_sum_daily_score REAL NOT NULL,
                        pl3_current_score REAL NOT NULL,
                        pl3_play_games INTEGER NOT NULL,
                        player4 TEXT NOT NULL,
                        pl4_sum_daily_score REAL NOT NULL,
                        pl4_current_score REAL NOT NULL,
                        pl4_play_games INTEGER NOT NULL,
                        score2 INTEGER NOT NULL,
                        win_lose2 INTEGER NOT NULL,
                        max_min2 INTEGER NOT NULL,
                        more10_2 INTEGER NOT NULL,
                        pair2_avr_score REAL NOT NULL,
                        pair2_Ea REAL NOT NULL,
                        pair2_Sa REAL NOT NULL,
                        pair2_Ra REAL NOT NULL);'''
    cursor = conn.cursor()
    cursor.execute(stat_table)
    conn.commit()
    cursor.fetchall()
    #print('Таблица статистики была создана заново')
    cursor.close()


#create_table()
def create_connection():
    """ create a database connection to the SQLite database
        specified by db_file
    :return: Connection object or None
    """
    db_file = 'list_of_games.db'
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)

    return conn

def drop_table_stat(table):

    conn = create_connection()
    sql = f"""DROP table IF EXISTS {table}"""

    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    #print('Таблица статистики была удалена')
    return cur.lastrowid

def count_games_in_day_from_db(date):
    #  counting how many games in special date already in DB table state
    conn = create_connection()
    sql = f"""SELECT count(*) FROM games WHERE date = '{date}'"""

    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    return cur.fetchone()

def del_games_from_db(date):
    conn = create_connection()
    sql = f"""DELETE FROM games WHERE date = '{date}'"""

    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    #print(f'!!!!   Все игры за {date} число удалены. Можете загружать повторно. ')
    return cur.lastrowid

def insert_games(task):
    """
    Create a new task
    :param conn:
    :param task:
    :return:
    """
    conn = create_connection()
    sql = ''' INSERT INTO games(date,player1,player2,player3,player4,score1,score2)
              VALUES(?,?,?,?,?,?,?) '''

    cur = conn.cursor()
    cur.execute(sql, task)
    conn.commit()
    return cur.lastrowid

def insert_stat(task):

    # create a database connection
    conn = create_connection()

    sql = ''' INSERT INTO stat(date,player1,pl1_sum_daily_score,pl1_current_score,pl1_play_games,
                                    player2,pl2_sum_daily_score,pl2_current_score,pl2_play_games,
                                    score1,win_lose1,max_min1,more10_1, 
                            pair1_avr_score,pair1_Ea,pair1_Sa,pair1_Ra,
                                    player3,pl3_sum_daily_score,pl3_current_score,pl3_play_games,
                                    player4,pl4_sum_daily_score,pl4_current_score,pl4_play_games,
                                    score2,win_lose2,max_min2,more10_2,
                            pair2_avr_score,pair2_Ea,pair2_Sa,pair2_Ra)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) '''

    cur = conn.cursor()
    cur.execute(sql, task)
    conn.commit()
    return cur.lastrowid

def insert_player(task):
    """
    Create a new task
    :param conn:
    :param task:
    :return:
    """
    conn = create_connection()
    sql = ''' INSERT INTO players(name,score,daily_score)
              VALUES(?,?,?) '''

    cur = conn.cursor()
    cur.execute(sql, task)
    conn.commit()
    return cur.lastrowid

def check_date(date):
    # create a database connection
    conn = create_connection()
    sql = ''' select date from games where date = (?) '''
    cur = conn.cursor()
    cur.execute(sql, (date,))
    result = cur.fetchall()
    if len(result) == 0:
        return True
    else:
        return False

def table_date_list():
    # create a database connection
    conn = create_connection()
    sql = ''' select distinct date from games order by date asc; '''
 #   sql = ''' select distinct date from games order by strftime('%s', date); '''
    conn.row_factory = lambda cursor, row: row[0]
    cur = conn.cursor()
    cur.execute(sql)
    result = cur.fetchall()
    return list(result)

def export_one_day_games(date):
    # create a database connection
    conn = create_connection()
    sql = ''' select player1,player2,player3,player4,score1,score2 from games where date = ? order by id asc ;'''
    conn.row_factory = sqlite3.Row
    values = conn.execute(sql, (date,)).fetchall()
    list = []
    for item in values:
        list.append({k: item[k] for k in item.keys()})
    return list

def create_view_table():
    conn = create_connection()
    sql = ''' create view if not exists v_player_stat as
select id, date, player1 player, pl1_current_score current_score, player2 partner, pair1_avr_score pair_avr_score, pair1_Ea Ea, pair1_Sa Sa, pair1_Ra Ra, win_lose1 win_lose, max_min1 max_min, more10_1 more10 from stat t
union all select id, date, player2 player, pl2_current_score, player1 partner, pair1_avr_score, pair1_Ea, pair1_Sa, pair1_Ra, win_lose1 win_lose, max_min1 max_min, more10_1 more10 from stat t
union all select id, date, player3 player, pl3_current_score, player4 partner, pair2_avr_score, pair2_Ea, pair2_Sa, pair2_Ra, win_lose2 win_lose, max_min2 max_min, more10_2 more10 from stat t
union all select id, date, player4 player, pl4_current_score, player3 partner, pair2_avr_score, pair2_Ea, pair2_Sa, pair2_Ra, win_lose2 win_lose, max_min2 max_min, more10_2 more10 from stat t;
'''

    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    return cur.lastrowid

def drop_view_table():
    conn = create_connection()
    sql = ''' drop view if exists v_player_stat;'''
    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    return cur.lastrowid

def select_stat1():

    ch_start = read_config('ch_start')
    ch_end = read_config('ch_end')

    Players_Championship_Day_Stat = '''
    select player   -- Игрок
     , 1.0*coalesce(sum(case when win_lose = 1 then 1 end), 0)/count(*) pct_win -- % побед
     , count(case when win_lose = 1 then 1 end) cnt_win -- количество побед
     , count(case when win_lose = 0 then 1 end) cnt_lost -- количество поражений
     , count(*) cnt_game -- количество всего игр 
     , sum(Ra) Ra --дельта счета за день
  from v_player_stat t1
  where t1.date = (select max(date) from stat)
  group by player
  order by pct_win desc, Ra desc; -- сортировка по процентам побед, дельте 
    '''

    Players_Championship_Total_Stat = '''
    with d as
( select distinct date from stat where date between ? and ? -- параметры, определяющие период турнира
), pl as
( select t1.player, min(t1.date) min_date from d inner join v_player_stat t1 on t1.date = d.date group by t1.player
), t2 as
( select pl.player
       , d.date
       , sum(count(win_lose)) over (partition by pl.player order by d.date) cnt_game
       , sum(sum(case when win_lose = 1 then 1 end)) over (partition by pl.player order by d.date) cnt_win
       , 1.0*coalesce(sum(sum(case when win_lose = 1 then 1 end)) over (partition by pl.player order by d.date), 0)/sum(count(win_lose)) over (partition by pl.player order by d.date) pct_win -- % побед
       , max(d.date) over () last_date
    from d cross join pl
         left outer join v_player_stat t1 on t1.date = d.date and t1.player = pl.player and d.date >= pl.min_date  
    group by pl.player, d.date
), t3 as
( select t2.*
       , rank() over (partition by date order by pct_win desc) player_rank
    from t2
), t4 as
( select t3.*
       , lag(player_rank) over (partition by player order by date) lag_player_rank
    from t3
)
select player
     , cnt_game -- сколько игр было сыграно
     , pct_win -- % побед
     , case when lag_player_rank <> player_rank then lag_player_rank-player_rank end diff_rank
     , 1 prize -- Добавить вхождение в призы (10 игроков/40% от avg)
  from t4
  where date = last_date
  order by pct_win desc, cnt_game desc;
    '''


    Players_Last_Day_Stat = '''select player
     , sum(Ra) last_date_Ra --дельта счета за этот день
     , count(case when win_lose = 1 then 1 end) cnt_win -- количество побед
     , count(case when win_lose = 0 then 1 end) cnt_lost -- количество поражений
     , count(case when win_lose = 1 and max_min = 1 then 1 end) cnt_tie_win -- количество больше-меньше побед
     , count(case when win_lose = 0 and max_min = 1 then 1 end) cnt_tie_lost -- количество больше-меньше поражений
     , count(*) cnt_game -- количество всего игр 
     , 100*coalesce(sum(case when win_lose = 1 then 1 end), 0)/count(*) pct_win -- % побед
  from v_player_stat t1
  where date = (select max(date) from stat)
  group by player
  order by sum(Ra) desc;
'''

    Players_Total_Stat = ''' with d as
( select distinct date from stat
), pl as
( select player, min(date) min_date from v_player_stat group by player
), t2 as
( select pl.player
       , d.date
       , sum(t1.Ra) last_date_Ra
       , sum(count(t1.Ra)) over (partition by pl.player order by d.date) cnt_game
       , max(case when sum(t1.Ra) is not null then d.date end) over (partition by pl.player) player_last_date
       , first_value(min(current_score)) over (partition by pl.player order by d.date) + sum(sum(t1.Ra)) over (partition by pl.player order by d.date) score
       , max(d.date) over () last_date
    from d inner join pl on d.date >= pl.min_date
         left outer join v_player_stat t1 on t1.date = d.date and t1.player = pl.player
    group by pl.player, d.date
), t3 as
( select t2.*
       , rank() over (partition by date order by score desc) player_rank
    from t2
), t4 as
( select t3.*
       , lag(player_rank) over (partition by player order by date) lag_player_rank
    from t3
)
select player
     , score -- текущий рейтинг
     , last_date_Ra -- дельту прироста за последнюю дату
     , cnt_game -- сколько игр было сыграно
     , player_last_date -- датой когда этот игрок играл последний раз
     , case when lag_player_rank <> player_rank then lag_player_rank-player_rank end diff_rank
  from t4
  where date = last_date
  order by score desc;'''

    Players_day_games_list_stat = '''select row_number() over (order by id), player1, pl1_current_score , pl1_play_games , 
                player2 , pl2_current_score , pl2_play_games , score1 , pair1_avr_score , pair1_Ea ,pair1_Sa , pair1_Ra ,
                player3, pl3_current_score , pl3_play_games , player4 , pl4_current_score , pl4_play_games , 
                score2 , pair2_avr_score , pair2_Ea ,pair2_Sa , pair2_Ra
                    from stat t1
                    where t1.date = (select max(date) from stat)
                    order by id'''

    Players_win_pair_stat = '''
    select player, partner
     , count(*)
     , coalesce(sum(case when win_lose = 1 then 1 end), 0) cnt_win
     , coalesce(sum(case when win_lose = 0 then 1 end), 0) cnt_lost
     , 100*coalesce(sum(case when win_lose = 1 then 1 end), 0)/count(*) pct_win
  from v_player_stat t1
  group by player, partner
  having count(*) > 2
  order by player, pct_win desc;'''

    Players_win_opponent_stat = '''with t1 as
    ( select id, date, player1 player, player3 opponent, sign(pair1_Ra) win from stat t
    union all select id, date, player1 player, player4 opponent, sign(pair1_Ra) win from stat t
    union all select id, date, player2 player, player3 opponent, sign(pair1_Ra) win from stat t
    union all select id, date, player2 player, player4 opponent, sign(pair1_Ra) win from stat t
    union all select id, date, player3 player, player1 opponent, sign(pair2_Ra) win from stat t
    union all select id, date, player3 player, player2 opponent, sign(pair2_Ra) win from stat t
    union all select id, date, player4 player, player1 opponent, sign(pair2_Ra) win from stat t
    union all select id, date, player4 player, player2 opponent, sign(pair2_Ra) win from stat t
    )
    select player, opponent
         , count(*)
         , coalesce(sum(case when win>0 then 1 end), 0) cnt_win
         , coalesce(sum(case when win<0 then 1 end), 0) cnt_lost
         , 100*coalesce(sum(case when win>0 then 1 end), 0)/count(*) pct_win
      from t1
      group by player, opponent
      having count(*) > 2
      order by player, pct_win desc;'''

    Players_daily_stat ='''with d as
( select distinct date from stat
), pl as
( select distinct player from v_player_stat
) 
select d.date
     ,pl.player
     , case when count(Ra) <> 0 then 500 + coalesce(sum(sum(Ra)) over (partition by pl.player order by d.date), 0) end score -- текущий рейтинг
  from d cross join pl
       left outer join v_player_stat t1 on t1.date = d.date and t1.player = pl.player
  group by pl.player, d.date
  order by pl.player, d.date;'''


    import xlsxwriter

    workbook = xlsxwriter.Workbook('Total_Stat.xlsx')
    conn = create_connection()
    c = conn.cursor()

    # ---------START------------- Excel FONT FORMAT LIST -------------START----------------
    def_fmt = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'bold': False
    })

    def_fmt_color = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'num_format': '[Blue]General;[Red]-General;General',
        'font_size': 14,
        'bg_color': '#E4E4E4'
    })

    bold_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'font_size': 14
    })

    head_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'bg_color': '#DEDEDE',
        'font_size': 12,
        'text_wrap': True
    })
    game_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'bg_color': '#F5D6C7',
        'font_size': 15
    })

    score_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 17,
        'border': 2
    })

    rating_fmt = workbook.add_format({
        'bold': False,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 16,
        'border': 2
    })
    # ---------END------------- Excel FONT FORMAT LIST --------------END---------------

    # -----------------------------------------------------------------------------------------------
    #  Description worksheet creation

    worksheet = workbook.add_worksheet('Описание')
    descr_file = open('description.txt', 'r', encoding='utf8')

    row = col = 0
    for line in descr_file:
        worksheet.write(row, col, line.strip("\n"))
        row += 1

    descr_file.close()

    # -----------------------------------------------------------------------------------------------
    # drop view table
    drop_view_table()
    #  Create view table
    create_view_table()

    # -----------------------------------------------------------------------------------------------
    #  ChampionShip Last Day worksheet creation
    worksheet = workbook.add_worksheet('Чемпион Дня')
    mysel = c.execute(Players_Championship_Day_Stat)
    header = ['Место', 'Игрок', 'Рейтинг', 'Победы', 'Поражения', 'Всего Игр', 'Дельта']
    for idx, col in enumerate(header):
        worksheet.write(0, idx, col, head_fmt)  # write the column name one time in a row

    # write all data from SELECT. keep 1 row and 1 column NULL
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            if j == 1:
                value = round(value, 3)
            elif j == 5:
                value = round(value, 0)
            worksheet.write(i + 1, j + 1, value, def_fmt)

    worksheet.write_column(1, 0, [i for i in range(1, len(c.execute(Players_Championship_Day_Stat).fetchall()) + 1)], head_fmt)  # make and insert column 1 with index

    worksheet.set_column('B:B', 18)
    worksheet.set_column('C:I', 11)

    # -----------------------------------------------------------------------------------------------
    worksheet = workbook.add_worksheet('Чемпионат 2023')
    mysel = c.execute(Players_Championship_Total_Stat, (ch_start, ch_end))
    header = ['Место', 'Игрок', 'Всего Игр', 'Рейтинг', 'Динамика']
    for idx, col in enumerate(header):
        worksheet.write(0, idx, col, head_fmt)  # write the column name one time in a row

    # write all data from SELECT. keep 1 row and 1 column NULL
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            if j == 3:
                worksheet.write(i + 1, j + 1, value, def_fmt_color)
            elif j == 0:
                worksheet.write(i + 1, j + 1, value, bold_fmt)

            elif j == 2:
                value = round(value, 3)
                worksheet.write(i + 1, j + 1, value, def_fmt)
            else:
                worksheet.write(i + 1, j + 1, value, def_fmt)

    worksheet.write_column(1, 0, [i for i in range(1, len(c.execute(Players_Championship_Total_Stat, (ch_start, ch_end)).fetchall()) + 1)],
                           head_fmt)  # make and insert column 1 with index

    worksheet.set_column('B:B', 18)
    worksheet.set_column('C:G', 11)
    worksheet.set_column('F:F', 18)

    # -----------------------------------------------------------------------------------------------
    #  Last-Day worksheet creation
    worksheet = workbook.add_worksheet('Рейтинг Дня')
    mysel = c.execute(Players_Last_Day_Stat)
    header = ['Место', 'Игрок', 'Дельта', 'Победы', 'Поражения', 'Win-Bal', 'Loss-Bal', 'Всего Игр', '% Побед']
    for idx, col in enumerate(header):
        worksheet.write(0, idx, col, head_fmt)  # write the column name one time in a row

    # write all data from SELECT. keep 1 row and 1 column NULL
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            if isinstance(value, float):
                value = int(value)
            worksheet.write(i + 1, j + 1, value, def_fmt)

    worksheet.write_column(1, 0, [i for i in range(1, len(c.execute(Players_Last_Day_Stat).fetchall()) + 1)], head_fmt)  # make and insert column 1 with index

    worksheet.set_column('B:B', 18)
    worksheet.set_column('C:I', 11)

    # -----------------------------------------------------------------------------------------------
    worksheet = workbook.add_worksheet('Общий Рейтинг')
    mysel=c.execute(Players_Total_Stat)
    header = ['Место', 'Игрок', 'Рейтинг', 'Дельта', 'Всего Игр', 'Последняя Игра', 'Динамика']
    for idx, col in enumerate(header):
        worksheet.write(0, idx, col, head_fmt)  # write the column name one time in a row

    # write all data from SELECT. keep 1 row and 1 column NULL
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            if j == 5:
                worksheet.write(i + 1, j + 1, value, def_fmt_color)
            elif j == 0:
                worksheet.write(i + 1, j + 1, value, bold_fmt)

            elif isinstance(value, float):
                value = int(value)
                worksheet.write(i + 1, j + 1, value, def_fmt)
            else:
                worksheet.write(i + 1, j + 1, value, def_fmt)

    worksheet.write_column(1, 0, [i for i in range(1, len(c.execute(Players_Total_Stat).fetchall()) + 1)], head_fmt)  # make and insert column 1 with index

    worksheet.set_column('B:B', 18)
    worksheet.set_column('C:G', 11)
    worksheet.set_column('F:F', 18)
    # -----------------------------------------------------------------------------------------------


    worksheet = workbook.add_worksheet(f'Все игры дня')
    mysel = c.execute(Players_day_games_list_stat)

    r = 1

    worksheet.set_column('A:A', 10)
    worksheet.set_column('C:I', 11)
    worksheet.set_column('B:B', 13)

    for i, row in enumerate(mysel):
        header = ['', 'Игрок', 'Рейтинг', 'Всего      Игр', 'Счет Партии', 'Средний Рейтинг', 'Ожидание (Ea)',
                  'Очки     (Sa)', 'Дельта     (Ra)']
        for idx, col in enumerate(header):
            worksheet.write(r - 1, idx, col, head_fmt)  # write the column name one time in a row
        worksheet.merge_range(r, 0, r + 1, 0, 'Пара 1', head_fmt)
        worksheet.merge_range(r + 2, 0, r + 3, 0, 'Пара 2', head_fmt)
        for j, value in enumerate(row):
            if j == 0:  # Counting Games
                worksheet.write(r - 1, 0, f'№ {value}', game_fmt)
            elif j == 12:  # Name Player 3
                r += 2
                worksheet.write(r, 1, value, bold_fmt)
            elif j == 1:  # Name Player 1
                worksheet.write(r, 1, value, bold_fmt)
            elif j == 2 or j == 13:  # Current score Player 1 and 3
                worksheet.write(r, 2, round(value), def_fmt)
            elif j == 3 or j == 14:  # Play games Player 1 and 3
                worksheet.write(r, 3, value, def_fmt)
            elif j == 4 or j == 15:  # Name Player 2 and 4
                worksheet.write(r + 1, 1, value, bold_fmt)
            elif j == 5 or j == 16:  # Current score Player 2 and 4
                worksheet.write(r + 1, 2, round(value, 1), def_fmt)
            elif j == 6 or j == 17:  # Play games Player 2 and 4
                worksheet.write(r + 1, 3, value, def_fmt)
            elif j == 7 or j == 18:  # Game score Pair 1 and Pair 2
                worksheet.merge_range(r, 4, r + 1, 4, value, score_fmt)
            elif j == 8 or j == 19:  # Average rating Pair 1 and Pair 2
                worksheet.merge_range(r, 5, r + 1, 5, round(value), rating_fmt)
            elif j == 9 or j == 20:  # Ea pair 1 and pair 2
                worksheet.merge_range(r, 6, r + 1, 6, value, def_fmt)
            elif j == 10 or j == 21:  # Sa pair 1 and pair 2
                worksheet.merge_range(r, 7, r + 1, 7, value, def_fmt)
            elif j == 11 or j == 22:  # Ra pair 1 and pair 2
                worksheet.merge_range(r, 8, r + 1, 8, round(value, 2), rating_fmt)
        r += 4

    # -----------------------------------------------------------------------------------------------
    worksheet = workbook.add_worksheet('Лучший Напарник')
    mysel=c.execute(Players_win_pair_stat)
    header = ['Игрок', 'Партнер', 'Всего Игр', 'Побед', 'Поражений', '% Побед']
    for idx, col in enumerate(header):
        worksheet.write(0, idx, col, head_fmt)  # write the column name one time in a row

    # write all data from SELECT. keep 1 row and 1 column NULL
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            if isinstance(value, float):
                value = int(value)
            worksheet.write(i + 1, j, value, def_fmt)

    # worksheet.write_column(1, 0, [i for i in range(1, len(c.execute(Players_win_pair_stat).fetchall()) + 1)])  # make and insert column 1 with index

    # here, we make both 1st column/row bold
    worksheet.set_column('A:B', 18)
    worksheet.set_column('C:F', 11)
    worksheet.autofilter(0, 0, 0, 0)

    # -----------------------------------------------------------------------------------------------
    worksheet = workbook.add_worksheet('Кому Проигрываешь')
    mysel=c.execute(Players_win_opponent_stat)
    header = ['Игрок', 'Соперник', 'Всего Игр', 'Побед', 'Поражений', '% Побед']
    for idx, col in enumerate(header):
        worksheet.write(0, idx, col, head_fmt)  # write the column name one time in a row

    # write all data from SELECT. keep 1 row and 1 column NULL
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            if isinstance(value, float):
                value = int(value)
            worksheet.write(i + 1, j, value, def_fmt)

    #worksheet.write_column(1, 0, [i for i in range(1, len(c.execute(Players_win_opponent_stat).fetchall()) + 1)])  # make and insert column 1 with index

    # here, we make both 1st column/row bold
    worksheet.set_column('A:B', 18)
    worksheet.set_column('C:F', 11)
    worksheet.autofilter(0, 0, 0, 0)

    #-----------------------------------------------------------------------------------------------
    worksheet = workbook.add_worksheet('График Рейтинга')
    mysel = c.execute(Players_daily_stat)

    r = 1
    cl = 0
    name = ''
    dt = ''
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            if j == 0:
                if i == 0:
                    dt = value
                    worksheet.write(r, 0, value)
                if value != dt:
                    worksheet.write(r, 0, value)
            elif j == 1:
                if value != name:
                    cl += 1
                    r = 1
                    name = value
                    worksheet.write(0, cl, value)
            elif j == 2:
                if isinstance(value, float):
                    value = int(value)
                    worksheet.write(r, cl, value)
        r += 1

    chart = workbook.add_chart({'type': 'line'})
    # Configure the series of the chart from the dataframe data.
    for i in range(cl):
        col = i + 1
        chart.add_series({
            'name': ['График Рейтинга', 0, col],
            'categories': ['График Рейтинга', 1, 0, r-1, 0],
            'values': ['График Рейтинга', 1, col, r-1, col],
        })

    # Configure the chart axes.
    chart.set_x_axis({'name': 'дата игры'})
    chart.set_y_axis({'name': 'Рейтинг'})
                      #'major_gridlines': {'visible': False}})
    chart.set_legend({'position': 'top'})
    chart.show_blanks_as('span')
    chart.set_size({'width': 1200, 'height': 580})
    # Insert the chart into the worksheet.
    worksheet.insert_chart('C3', chart)
    # -----------------------------------------------------------------------------------------------

    workbook.close()


def select_stat2():
    Players_daily_stat = '''select row_number() over (order by id), player1, pl1_current_score , pl1_play_games , 
            player2 , pl2_current_score , pl2_play_games , score1 , pair1_avr_score , pair1_Ea ,pair1_Sa , pair1_Ra ,
            player3, pl3_current_score , pl3_play_games , player4 , pl4_current_score , pl4_play_games , 
            score2 , pair2_avr_score , pair2_Ea ,pair2_Sa , pair2_Ra
                from stat t1
                where t1.date = (select max(date) from stat)
                order by id'''

    from xlsxwriter.workbook import Workbook

    workbook = Workbook('Total_Stat_test.xlsx')
    worksheet = workbook.add_worksheet('1')

    conn = create_connection()
    c = conn.cursor()
    mysel = c.execute(Players_daily_stat)

    r = 1
    bold_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'font_size': 14,
    })
    head_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'bg_color': '#DEDEDE',
        'font_size': 12,
        'text_wrap': True
    })

    game_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2,
        'bg_color': '#F5D6C7',
        'font_size': 15
    })

    def_fmt = workbook.add_format({
        'bold': False,
        'align': 'center',
        'valign': 'vcenter',
        'border': 2
    })
    score_fmt = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 17,
        'border': 2
    })
    rating_fmt = workbook.add_format({
        'bold': False,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 16,
        'border': 2
    })
    worksheet.set_column('A:A', 10)
    worksheet.set_column('C:I', 11)
    worksheet.set_column('B:B', 13)

    for i, row in enumerate(mysel):
        header = ['','Игрок', 'Рейтинг', 'Всего      Игр', 'Счет Партии', 'Средний Рейтинг', 'Ожидание (Ea)', 'Очки     (Sa)', 'Дельта     (Ra)']
        for idx, col in enumerate(header):
            worksheet.write(r-1, idx, col, head_fmt)  # write the column name one time in a row
        worksheet.merge_range(r, 0, r + 1, 0, 'Пара 1', head_fmt)
        worksheet.merge_range(r+2, 0, r + 3, 0, 'Пара 2', head_fmt)
        for j, value in enumerate(row):
            if j == 0:               #  Counting Games
                worksheet.write(r - 1, 0, f'№ {value}', game_fmt)
            elif j == 12:            #  Name Player 3
                r += 2
                worksheet.write(r, 1, value, bold_fmt)
            elif j == 1:             #  Name Player 1
                worksheet.write(r, 1, value, bold_fmt)
            elif j == 2 or j == 13:  #  Current score Player 1 and 3
                worksheet.write(r, 2, round(value), def_fmt)
            elif j == 3 or j == 14:  #  Play games Player 1 and 3
                worksheet.write(r, 3, value, def_fmt)
            elif j == 4 or j == 15:  #  Name Player 2 and 4
                worksheet.write(r+1, 1, value, bold_fmt)
            elif j == 5 or j == 16:  #  Current score Player 2 and 4
                worksheet.write(r+1, 2, round(value, 1), def_fmt)
            elif j == 6 or j == 17:  #  Play games Player 2 and 4
                worksheet.write(r+1, 3, value, def_fmt)
            elif j == 7 or j == 18:  #  Game score Pair 1 and Pair 2
                worksheet.merge_range(r, 4, r + 1, 4, value, score_fmt)
            elif j == 8 or j == 19:  #  Average rating Pair 1 and Pair 2
                worksheet.merge_range(r, 5, r + 1, 5, round(value), rating_fmt)
            elif j == 9 or j == 20:  #  Ea pair 1 and pair 2
                worksheet.merge_range(r, 6, r+1, 6, value, def_fmt)
            elif j == 10 or j == 21:  #  Sa pair 1 and pair 2
                worksheet.merge_range(r, 7, r+1, 7, value, def_fmt)
            elif j == 11 or j == 22:  # Ra pair 1 and pair 2
                worksheet.merge_range(r, 8, r + 1, 8, round(value, 2), rating_fmt)
        r += 4

    # chart = workbook.add_chart({'type': 'line'})
    # # Configure the series of the chart from the dataframe data.
    # for i in range(cl):
    #     col = i + 1
    #     chart.add_series({
    #         'name': ['Daily', 0, col],
    #         'categories': ['Daily', 1, 0, r, 0],
    #         'values': ['Daily', 1, col, r, col],
    #     })
    #
    # # Configure the chart axes.
    # chart.set_x_axis({'name': 'Date'})
    # chart.set_y_axis({'name': 'Value', 'major_gridlines': {'visible': False}})
    #
    # # Insert the chart into the worksheet.
    # worksheet.insert_chart('D3', chart)

    workbook.close()


#select_stat2()


# class Stat_to_Excel:
#
#     workbook = Workbook('Total_Stat.xlsx')
#     db_file = 'list_of_games.db'
#
#     Players_Last_Day_Stat = '''with t1 as
#     ( select id, date, player1 player, player2 partner, pair1_avr_score pair_avr_score, pair1_Ea Ea, pair1_Sa Sa, pair1_Ra Ra, sign(pair1_Ra) win, sign(case when pair1_Ra > -1 and pair1_Ra < 1 then pair1_Ra end) tie_win from stat t
#     union all select id, date, player2 player, player1 partner, pair1_avr_score, pair1_Ea, pair1_Sa, pair1_Ra, sign(pair1_Ra) win, sign(case when pair1_Ra > -1 and pair1_Ra < 1 then pair1_Ra end) tie_win from stat t
#     union all select id, date, player3 player, player4 partner, pair2_avr_score, pair2_Ea, pair2_Sa, pair2_Ra, sign(pair2_Ra) win, sign(case when pair1_Ra > -1 and pair1_Ra < 1 then pair1_Ra end) tie_win from stat t
#     union all select id, date, player4 player, player3 partner, pair2_avr_score, pair2_Ea, pair2_Sa, pair2_Ra, sign(pair2_Ra) win, sign(case when pair1_Ra > -1 and pair1_Ra < 1 then pair1_Ra end) tie_win from stat t
#     )
#     select player                                                -- игрок
#          , sum(Ra) last_date_Ra                                  -- дельта счета за этот день
#          , count(case when win = 1 then 1 end) cnt_win           -- количество побед
#          , count(case when win = -1 then 1 end) cnt_lost         -- количество поражений
#          , count(case when tie_win = 1 then 1 end) cnt_tie_win   -- количество больше-меньше побед
#          , count(case when tie_win = -1 then 1 end) cnt_tie_lost -- количество больше-меньше поражений
#          , count(*) cnt_game                                     -- количество всего игр
#       from t1
#       where date = (select max(date) from stat)
#       group by player
#       order by sum(Ra) desc;'''
#
#     Players_Total_Stat = ''' with t1 as
#     ( select id, date, player1 player, player2 partner, pair1_avr_score pair_avr_score, pair1_Ea Ea, pair1_Sa Sa, pair1_Ra Ra from stat t
#     union all select id, date, player2 player, player1 partner, pair1_avr_score, pair1_Ea, pair1_Sa, pair1_Ra from stat t
#     union all select id, date, player3 player, player4 partner, pair2_avr_score, pair2_Ea, pair2_Sa, pair2_Ra from stat t
#     union all select id, date, player4 player, player3 partner, pair2_avr_score, pair2_Ea, pair2_Sa, pair2_Ra from stat t
#     ), t2 as
#     ( select t1.*
#            , rank() over (partition by player order by date desc) player_date_rank
#            , rank() over (order by date desc) date_rank
#         from t1
#     )
#     select player                                                -- игрок
#          , 500 + sum(Ra) score                                   -- текущий рейтинг
#          , sum(case when date_rank = 1 then Ra end) last_date_Ra -- дельту прироста за последнюю дату
#          , count(*) cnt_game                                     -- сколько игр было сыграно
#          , max(date) last_date                                   -- дата когда этот игрок играл последний раз
#       from t2
#       group by player
#       order by sum(Ra) desc;'''
#
#     def create_connection():
#     """ create a database connection to the SQLite database
#         specified by db_file
#     :return: Connection object or None
#     """
#     db_file = self.db_file
#     conn = None
#     try:
#         conn = sqlite3.connect(db_file)
#     except Error as e:
#         print(e)
#
#     return conn

#     def export_report(self, report, header):
#
#         workbook = self.workbook
#         conn = create_connection()
#         c = conn.cursor()
#
#         worksheet = workbook.add_worksheet('Last-Day')
#         mysel = c.execute(Players_Last_Day_Stat)
#         header = ['Место', 'Игрок', 'Дельта', 'Победы', 'Поражения', 'Win(max/min)', 'Loss(max/min)', 'Всего Игр']
#         for idx, col in enumerate(header):
#             worksheet.write(0, idx, col)  # write the column name one time in a row
#
#         # write all data from SELECT. keep 1 row and 1 column NULL
#         for i, row in enumerate(mysel):
#             for j, value in enumerate(row):
#                 if isinstance(value, float):
#                     value = int(value)
#                 worksheet.write(i + 1, j + 1, value)
#
#         worksheet.write_column(1, 0, [i for i in range(1, len(c.execute(
#             Players_Last_Day_Stat).fetchall()) + 1)])  # make and insert column 1 with index
#
#         # here, we make both 1st column/row bold
#         bold_fmt = workbook.add_format({'bold': True})
#         worksheet.set_row(0, None, bold_fmt)
#         worksheet.set_column(0, 0, None, bold_fmt)